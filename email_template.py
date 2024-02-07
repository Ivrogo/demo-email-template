import gspread
from google.auth.transport.requests import Request
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from openpyxl import load_workbook

def obtain_api_credentials():
    API_scopes = [
        
        "https://www.googleapis.com/auth/gmail.settings.basic",
        "https://www.googleapis.com/auth/gmail.settings.sharing"
        
    ]
    credentials = Credentials.from_service_account_file("credentials.json", scopes=API_scopes)
    return credentials


def get_emails_from_excel(excel_path):
    wb = load_workbook(excel_path)
    sheet = wb.active
    
    emails_from_excel = [row[3] for row in sheet.iter_rows(min_row=2, values_only=True)]
    
    return emails_from_excel


def get_user_info_from_excel(email, excel_path):
    wb = load_workbook(excel_path)
    sheet = wb.active
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(f'Checking row: {row}')
        if row[3] == email:
            print(f'Found matching email: {email}')
            return row[0], row[1], row[2]
    
    return None



def update_gmail_signature(email, signature, live=False): 
    
    creds = obtain_api_credentials()
    if live:
        creds_delegated = creds.with_subject(email)
    
        #Creamos la instancia de la API de gmail
        gmail_service = build('gmail', 'v1', credentials=creds_delegated)
        addresses = gmail_service.users().settings().sendAs().list(userId="me", fields="seendAs(isPrimary, sendAsEmail)").execute().get("sendAs")
        
        address = None
        for address in addresses:
            if address.get('isPrimary'):
                break
        if address:
            rsp = gmail_service.users().settings().sendAs().patch(userId='me', sendAsEmail=address['sendAsEmail'], body={'signature': signature}).execute()
            print(f'Signature changed for: {email}')
            print(f'New signature:\n{signature}')
        else:
            print(f'Could not find primary address for: {email}')
 
           

with open('template.html', 'r') as file:
    template_html = file.read()
    
excel_path = './datos_equipo_de_trabajo_demomails.xlsx'


usuarios = [{'email': email} for email in get_emails_from_excel(excel_path)]

for user in usuarios:
    email = user['email']
    user_info = get_user_info_from_excel(email, excel_path)
    
    if user_info:
        departamento, cargo, nombre = user_info
        if nombre and departamento and cargo:
            firma_personalizada = template_html.replace('{{NOMBRE}}', nombre).replace('{{DEPARTAMENTO}}', departamento).replace('{{CARGO}}', cargo)
            print(firma_personalizada)
            update_gmail_signature(email, firma_personalizada)
        else:
            break
    else:
        print(f'No se encontro información para {email} en el archivo Excel')
    
print('Proceso completado')
    